import concurrent.futures
import glob
import json
import logging
import os
import os.path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_single_xlsx_file(instruction_id: List[str], en_utt: List[str], en_annotation: List[str],
                              file: str,
                              destination_folder: str):
    logging.info(f"Parsing the file {file}")
    xx_utterance = []
    xx_annot = []
    lang = ""
    xx_file = open(file, "r", encoding='utf-8')
    for line in xx_file:
        instruction = json.loads(line)
        utter = instruction.get("utt")
        annot = instruction.get("annot_utt")
        lang = instruction.get("locale")
        xx_utterance.append(utter)
        xx_annot.append(annot)
    xx_file.close()

    separator = '-'
    language = lang.split(separator, 1)[0]

    df = pd.DataFrame({"ID": instruction_id, "English Utterance": en_utt, "English Annotated": en_annotation,
                       "" + language + " Utterance": xx_utterance, "" + language + " Annotated": xx_annot})

    df.to_excel(f"{destination_folder}/en-" + language + ".xlsx", index=False)


def generate_xlsx_files(path_to_data: str, destination_folder: str) -> None:
    instruction_id = []
    en_utterance = []
    en_annot = []
    massive_dataset_files = list(glob.iglob(f'{path_to_data}/*.jsonl'))

    print(path_to_data)
    english_file = open(f"{path_to_data}/en-US.jsonl", "r", encoding='utf-8')
    for line in english_file:
        instruction = json.loads(line)
        utter = instruction.get("utt")
        inst_id = instruction.get("id")
        annot = instruction.get("annot_utt")
        instruction_id.append(inst_id)
        en_utterance.append(utter)
        en_annot.append(annot)

    os.makedirs(destination_folder, exist_ok=True)

    with concurrent.futures.ProcessPoolExecutor() as p:
       for file in massive_dataset_files:
            p.submit(generate_single_xlsx_file, *(instruction_id, en_utterance,
                                                  en_annot,
                                                  file,
                                                  destination_folder))

    logging.info(f"Task successful! Your en-xx.xlsx files have been generated in the folder: {destination_folder}")


def generate_xlsx_sheets(path_to_data: str, destination_file: str) -> None:
    if os.path.isfile(destination_file):
        instruction_id = []
        en_utterance = []
        en_annot = []
        massive_dataset_files = list(glob.iglob(f'{path_to_data}/*.jsonl'))

        english_file = open(f"{path_to_data}/en-US.jsonl", "r", encoding='utf-8')
        for line in english_file:
            instruction = json.loads(line)
            utter = instruction.get("utt")
            inst_id = instruction.get("id")
            annot = instruction.get("annot_utt")
            instruction_id.append(inst_id)
            en_utterance.append(utter)
            en_annot.append(annot)

        for file in massive_dataset_files:
            xx_utterance = []
            xx_annot = []
            lang = ""
            xx_file = open(file, "r", encoding='utf-8')
            for line in xx_file:
                instruction = json.loads(line)
                utter = instruction.get("utt")
                annot = instruction.get("annot_utt")
                lang = instruction.get("locale")
                xx_utterance.append(utter)
                xx_annot.append(annot)
            xx_file.close()

            separator = '-'
            language = lang.split(separator, 1)[0]

            df = pd.DataFrame({"ID": instruction_id, "English Utterance": en_utterance, "English Annotated": en_annot,
                               "" + language + " Utterance": xx_utterance, "" + language + " Annotated": xx_annot})

            book = load_workbook(destination_file)
            book.create_sheet(f"en-{language}")
            b = book[f"en-{language}"]
            for r in dataframe_to_rows(df, index=False):
                b.append(r)
            book.save(destination_file)
        logging.info(f"Task successful! Your .xlsx file with separate en-xx sheets "
                     f"has been generated in the specified file: {destination_file}")
    else:
        logging.error('The xlsx file you have provided does not exist. Kindly provide an already existing .xlsx file')


def specific_lang_xlsx_file(path_to_data: str, name_of_lang: str, path_to_destination_folder: str) -> None:
    instruction_id = []
    en_utterance = []
    en_annot = []
    xx_utterance = []
    xx_annot = []
    lang = ""

    english_file = open(f"{path_to_data}/en-US.jsonl", "r", encoding='utf-8')
    for line in english_file:
        instruction = json.loads(line)
        utter = instruction.get("utt")
        inst_id = instruction.get("id")
        annot = instruction.get("annot_utt")
        instruction_id.append(inst_id)
        en_utterance.append(utter)
        en_annot.append(annot)

    path = f"{path_to_data}/{name_of_lang}.jsonl"

    if os.path.isfile(path):
        xx_file = open(path, "r", encoding='utf-8')
        for xx_line in xx_file:
            xx_instruction = json.loads(xx_line)
            utter = xx_instruction.get("utt")
            annot = xx_instruction.get("annot_utt")
            lang = xx_instruction.get("locale")
            xx_utterance.append(utter)
            xx_annot.append(annot)
        xx_file.close()

        separator = '-'
        language = lang.split(separator, 1)[0]

        df = pd.DataFrame({"ID": instruction_id, "English Utterance": en_utterance, "English Annotated": en_annot,
                           "" + language + " Utterance": xx_utterance, "" + language + " Annotated": xx_annot})
        path = path_to_destination_folder + r"\en-" + language + ".xlsx"
        if os.path.isdir(path_to_destination_folder):
            df.to_excel(path, index=False)
        else:
            os.mkdir(path_to_destination_folder)
            df.to_excel(path, index=False)

        logging.info(f"Task successful! Your en-xx.xlsx file for the language: {name_of_lang} has "
                     f"been created in the folder: {path_to_destination_folder}")
    else:
        logging.error(f"The language you have specified does not exist in the dataset."
                      f" Kindly check your format or spelling.")
